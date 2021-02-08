from openpyxl import Workbook
wb = Workbook()

ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "66ccff" # RGB 형태의 값으로 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
# Sheet, MySheet, YourSheet 이 순서대로

ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index 에 Sheet 생성(3번째로)

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근이 가능

print(wb.sheetnames) # Sheet 의 이름 순서대로 정렬

# Sheet 복사
new_ws["A1"] = "Test" # A1 은 1행 1렬. 여기에 Test 입력
# 이제 이 A1 의 데이터를 복사해볼게

target = wb.copy_worksheet(new_ws) # target 에 new_ws 즉 A1 의 값을 붙혀넣음
target.title = "Copied Sheet" # 그 target 의 이름을 이렇게 해볼게

wb.save("sample.xlsx")