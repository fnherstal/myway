from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# A1 셀에 1 이라는 값을 입력
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # A1 셀의 정보를 출력 (그냥 객체 정보가 뜸)
print(ws["A1"].value) # A1 셀의 '값'을 출력
print(ws["A10"].value) # 값이 없으면 'None' 을 출력

# 행 그리고 열을 적어도 A1 과 똑같은 값을 의미한다. 또 다른 타이핑 방식.
ws.cell(row=1, column=1)
# row = 1, 2, 3, . . .
# column = A(1), B(2), C(3), . . . 열을 이렇게 숫자로 적는거다
print(ws.cell(row=1,column=1).value) # ws["A1"].value
print(ws.cell(row=1,column=2).value) # ws["B1"].value
# 행과 열의 입력은 수학처럼 하되, 엑셀에서의 모습은 반대로 나온다
print(ws.cell(column=2, row=1).value) # B1 을 이렇게 열 행으로 적어도 괜찮다.

c = ws.cell(column=3, row=1, value=10) # ws["C1"] = 10
print(c.value) # ws["C1"]

from random import *
index = 1
# 반복문을 이용해서 랜덤 숫자를 넣기
for x in range(1, 11): # 10 개 row
    for y in range(1, 11): # 10 개 column
        #ws.cell(row = x, column = y, value = randint(0, 100)) # 0~100
        ws.cell(row = x, column = y, value = index) # 1 부터 더해가기
        index += 1

wb.save("sample.xlsx")