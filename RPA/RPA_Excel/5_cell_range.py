from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# 1 줄씩 데이터 넣기
ws.append(["번호", "영어", "수학"])
for i in range(1, 11): # 10 개 데이터 넣기
    ws.append([i, randint(0, 100), randint(0, 100)]) # 번호 영점수 수점수

col_B = ws["B"] # 영어 column 만 가져오기
# print(col_B)
# for cell in col_B:
#     print(cell.value)

col_range = ws["B:C"] # 영, 수 column 함께 가져오기
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

row_title = ws[1] # 1 번째 row 만 가지고 오기
# for cell in row_title:
#     print(cell.value)

row_range = ws[2:6] # 2번째 줄에서 6번째 줄까지 가져오기
# 원래는 2~5 인데 여기선 2~6 이다. 주의. 1~5번 학생.

# for rows in row_range:
#     for cell in rows:
#         print(cell.value, end=" ")
#     print() # 줄 바꿈

# Data 가 어떤 cell 에 있는지 알게 해줌
# from openpyxl.utils.cell import coordinate_from_string
#
# row_range = ws[2:ws.max_row] # 2번째 줄부터 마지막 줄까지
# for rows in row_range:
#     for cell in rows:
#         # print(cell.value, end=" ")
#         # print(cell.coordinate, end=" ") # A10, AZ250
#         xy= coordinate_from_string(cell.coordinate)
#         # print(xy, end=" ")
#         print(xy[0], end="") # A 등
#         print(xy[1], end=" ") # 1 등
#     print()

# 전체 rows
# print(tuple(ws.rows)) # 한 줄씩 튜플로
# for row in tuple(ws.rows):
#     print(row[2].value) # 수학 점수만

# 전체 columns
# print(tuple(ws.columns)) # 한 열씩 튜플로
# for column in tuple(ws.columns):
#     print(column[0].value)

# for row in ws.iter_rows(): # 전체 row
#     print(row[1].value) # 영어

# for column in ws.iter_cols(): # 전체 column
#     print(column[0].value)

# # 2~11 줄 까지, 2~3 열 까지 좌우좌우로
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     # print(row[0].value, row[1].value) # 영어
#     print(row)

# 상하상하로
for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col)

wb.save("sample.xlsx")